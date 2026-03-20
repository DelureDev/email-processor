#!/usr/bin/env python3
"""One-time: remove today's records from master so re-run produces a fresh report."""
import openpyxl
import shutil
import sqlite3

# 1. Backup and clean master.xlsx
path = './output/master.xlsx'
shutil.copy2(path, path.replace('.xlsx', '_pre_rerun.xlsx'))
wb = openpyxl.load_workbook(path)
ws = wb['Данные']
deleted = 0
for r in range(ws.max_row, 1, -1):
    date_col = ws.cell(row=r, column=11).value
    if str(date_col or '') == '20.03.2026':
        ws.delete_rows(r, 1)
        deleted += 1
wb.save(path)
wb.close()
print(f'Removed {deleted} rows from master.xlsx')

# 2. Clean master.csv backup too
import csv, os
csv_path = './output/master.csv'
if os.path.exists(csv_path):
    with open(csv_path, 'r', encoding='utf-8-sig', newline='') as f:
        rows = list(csv.reader(f, delimiter=';'))
    header = rows[0] if rows else []
    date_idx = header.index('Дата обработки') if 'Дата обработки' in header else -1
    if date_idx >= 0:
        kept = [header] + [r for r in rows[1:] if len(r) <= date_idx or r[date_idx] != '20.03.2026']
        removed_csv = len(rows) - len(kept)
        with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
            w = csv.writer(f, delimiter=';', lineterminator='\r\n')
            w.writerows(kept)
        print(f'Removed {removed_csv} rows from master.csv')

# 3. Clear today's processed IDs
conn = sqlite3.connect('processed_ids.db')
d = conn.execute("DELETE FROM processed_ids WHERE seen_at >= '2026-03-20'").rowcount
conn.commit()
conn.close()
print(f'Cleared {d} processed IDs')

print('\nReady. Run: python3 main.py')
