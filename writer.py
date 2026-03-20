"""
Writer — appends normalized records to master xlsx file on network drive.
"""
import os
import sys
import shutil
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from contextlib import contextmanager
from datetime import datetime
import logging
from parsers.utils import clean_dedup_val, norm_date_pad

logger = logging.getLogger(__name__)


@contextmanager
def _master_lock(master_path: str):
    """Exclusive file lock around master.xlsx writes. No-op on Windows (dev only)."""
    if sys.platform == 'win32':
        yield
        return
    import fcntl
    lock_path = master_path + '.lock'
    with open(lock_path, 'w') as lf:
        fcntl.flock(lf, fcntl.LOCK_EX)
        try:
            yield
        finally:
            fcntl.flock(lf, fcntl.LOCK_UN)


COLUMNS = ['ФИО', 'Дата рождения', '№ полиса', 'Начало обслуживания', 'Конец обслуживания', 'Страховая компания', 'Страхователь', 'Клиника', 'Комментарий в полис', 'Источник файла', 'Дата обработки']


def load_existing_keys(master_path: str) -> set:
    """Load dedup keys (ФИО + полис + начало + конец + клиника) from existing master file."""
    keys = set()
    if not os.path.exists(master_path):
        return keys
    try:
        dedup_cols = ['ФИО', '№ полиса', 'Начало обслуживания', 'Конец обслуживания', 'Клиника']
        header_cols = set(pd.read_excel(master_path, nrows=0).columns)
        available = dedup_cols if all(c in header_cols for c in dedup_cols) else dedup_cols[:-1]
        df = pd.read_excel(master_path, usecols=available)
        if 'Клиника' not in df.columns:
            df['Клиника'] = ''

        for col in available:
            df[col] = df[col].map(lambda v: clean_dedup_val(v))
        df['ФИО'] = df['ФИО'].str.upper().str.replace('Ё', 'Е', regex=False)
        for col in ['Начало обслуживания', 'Конец обслуживания']:
            df[col] = df[col].map(norm_date_pad)

        keys = set(zip(df['ФИО'], df['№ полиса'], df['Начало обслуживания'], df['Конец обслуживания'], df['Клиника']))
    except Exception as e:
        logger.error(f"Error loading existing keys: {e}")
    return keys

HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='2F5496')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
DATA_FONT = Font(name='Arial', size=10)
DATA_ALIGNMENT = Alignment(vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
COLUMN_WIDTHS = {
    'ФИО': 35,
    'Дата рождения': 16,
    '№ полиса': 25,
    'Начало обслуживания': 20,
    'Конец обслуживания': 20,
    'Страховая компания': 25,
    'Страхователь': 25,
    'Клиника': 25,
    'Комментарий в полис': 50,
    'Источник файла': 30,
    'Дата обработки': 16,
}


def write_to_master(records: list[dict], master_path: str, source_filename: str = ""):
    """Append records to master xlsx file. Creates file if it doesn't exist."""
    write_batch_to_master([(records, source_filename)], master_path)


def write_batch_to_master(batch: list[tuple[list[dict], str]], master_path: str):
    """
    Write multiple batches of records in a single open/save cycle.
    batch: list of (records, source_filename) tuples.
    Opens the workbook once regardless of how many files are in the batch.
    """
    if not batch:
        return

    os.makedirs(os.path.dirname(master_path) or '.', exist_ok=True)

    now = datetime.now().strftime('%d.%m.%Y')
    all_records = [
        {**r, 'Источник файла': source_filename, 'Дата обработки': now}
        for records, source_filename in batch
        for r in records
    ]

    with _master_lock(master_path):
        if os.path.exists(master_path):
            bak_path = master_path + '.bak'
            bak_created = False
            try:
                shutil.copy2(master_path, bak_path)
                bak_created = True
            except OSError as e:
                logger.warning(f"Could not create backup of master: {e}")
            try:
                _append_to_existing(all_records, master_path)
            except Exception as e:
                logger.error(f"Write to master failed: {e}")
                if bak_created:
                    try:
                        shutil.copy2(bak_path, master_path)
                        logger.warning(f"Restored master from backup after write failure")
                    except OSError as restore_err:
                        logger.error(f"Restore from backup also failed: {restore_err}")
                raise
            # Clean up backup after successful write
            if bak_created:
                try:
                    os.remove(bak_path)
                except OSError:
                    pass
        else:
            _create_new(all_records, master_path)

    logger.info(f"Wrote {len(all_records)} records ({len(batch)} files) to {master_path}")
    _export_csv(master_path, all_records)


def _export_csv(master_path: str, new_records: list[dict]) -> None:
    """Append new records to master CSV alongside master.xlsx (UTF-8 BOM, incremental)."""
    import csv as csv_mod
    csv_path = os.path.splitext(master_path)[0] + '.csv'
    try:
        write_header = not os.path.exists(csv_path)
        # Use utf-8-sig only for new files (writes BOM); utf-8 for appends (no mid-file BOM)
        encoding = 'utf-8-sig' if write_header else 'utf-8'
        with open(csv_path, 'a', newline='', encoding=encoding) as f:
            writer = csv_mod.DictWriter(f, fieldnames=COLUMNS, extrasaction='ignore', delimiter=';')
            if write_header:
                writer.writeheader()
            for record in new_records:
                writer.writerow({k: _safe(v) for k, v in record.items()})
        logger.info(f"CSV backup updated: {csv_path} (+{len(new_records)} rows)")
    except Exception as e:
        logger.warning(f"CSV backup failed: {e}")


def _safe(value) -> object:
    """Prevent formula injection by prefixing formula-like strings with apostrophe.
    Does not prefix negative numbers (e.g. -500 stays as-is).
    """
    if value is None:
        return ''
    s = str(value)
    if not s:
        return value
    c = s[0]
    if c in ('=', '+', '@', '\t', '\r', '|'):
        return "'" + s
    if c == '-' and (len(s) < 2 or not s[1].isdigit()):
        return "'" + s
    return value


def _migrate_xlsx_columns(ws):
    """Migrate old-layout worksheet by inserting Клиника and Комментарий в полис columns.
    Old layout: ФИО|ДР|Полис|Начало|Конец|СК|Страхователь|Источник|Дата обр
    New layout: ...same...|Клиника|Комментарий|Источник|Дата обр
    Shifts Источник файла and Дата обработки right by 2 to make room."""
    # Insert two columns before 'Источник файла' (was col 8, now cols 8-9 are new)
    ws.insert_cols(8, 2)
    # Set new headers
    ws.cell(row=1, column=8, value='Клиника').font = HEADER_FONT
    ws.cell(row=1, column=8).fill = HEADER_FILL
    ws.cell(row=1, column=8).alignment = HEADER_ALIGNMENT
    ws.cell(row=1, column=8).border = THIN_BORDER
    ws.cell(row=1, column=9, value='Комментарий в полис').font = HEADER_FONT
    ws.cell(row=1, column=9).fill = HEADER_FILL
    ws.cell(row=1, column=9).alignment = HEADER_ALIGNMENT
    ws.cell(row=1, column=9).border = THIN_BORDER
    # Set column widths for new columns
    ws.column_dimensions[get_column_letter(8)].width = COLUMN_WIDTHS.get('Клиника', 20)
    ws.column_dimensions[get_column_letter(9)].width = COLUMN_WIDTHS.get('Комментарий в полис', 20)


def _populate_styled_worksheet(ws, records: list[dict]):
    """Populate a worksheet with styled headers and data rows."""
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS.get(col_name, 20)

    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    for row_idx, record in enumerate(records, 2):
        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_safe(record.get(col_name, '')))
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = DATA_ALIGNMENT

    ws.freeze_panes = 'A2'


def build_styled_xlsx_bytes(records: list[dict]) -> bytes:
    """Build styled xlsx from records list, returns bytes. Used by notifier."""
    import io
    wb = Workbook()
    try:
        ws = wb.active
        ws.title = "Данные"
        _populate_styled_worksheet(ws, records)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    finally:
        wb.close()


def _create_new(records: list[dict], path: str):
    wb = Workbook()
    try:
        ws = wb.active
        ws.title = "Данные"
        _populate_styled_worksheet(ws, records)
        wb.save(path)
    finally:
        wb.close()


def _append_to_existing(records: list[dict], path: str):
    wb = load_workbook(path)
    try:
        if 'Данные' not in wb.sheetnames:
            raise ValueError(f"Sheet 'Данные' not found in {path}. Available sheets: {wb.sheetnames}")
        ws = wb['Данные']

        # Validate column order — auto-migrate if old layout is detected
        existing_headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        if existing_headers != COLUMNS:
            # Check if this is the known old layout (missing Клиника, Комментарий в полис)
            old_layout = ['ФИО', 'Дата рождения', '№ полиса', 'Начало обслуживания',
                          'Конец обслуживания', 'Страховая компания', 'Страхователь',
                          'Источник файла', 'Дата обработки']
            actual_non_none = [h for h in existing_headers if h is not None]
            if actual_non_none == old_layout:
                logger.info(f"Migrating {path} columns from old layout → adding Клиника, Комментарий в полис")
                _migrate_xlsx_columns(ws)
            else:
                logger.warning(f"Column mismatch in {path}: expected {COLUMNS}, got {existing_headers}")

        # Find actual last data row (max_row may include empty styled rows)
        next_row = ws.max_row + 1
        for r in range(ws.max_row, 0, -1):
            if any(ws.cell(row=r, column=c).value is not None for c in range(1, len(COLUMNS) + 1)):
                next_row = r + 1
                break

        for row_idx, record in enumerate(records, next_row):
            for col_idx, col_name in enumerate(COLUMNS, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=_safe(record.get(col_name, '')))
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                cell.alignment = DATA_ALIGNMENT

        # Update autofilter range
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"

        wb.save(path)
    finally:
        wb.close()
