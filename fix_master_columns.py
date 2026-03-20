#!/usr/bin/env python3
"""
One-time fix for master.xlsx after v1.9.2 wrote new-layout data into old-layout file.

What it does:
1. Removes 26 phantom rows re-ingested from own report email (Источник = '2042_records_2026-03-19.xlsx')
2. Fixes column layout: inserts Клиника + Комментарий в полис columns
   - Old rows: get empty Клиника/Комментарий, Источник+Дата shifted right
   - Today's rows: already have data in new order, just need header alignment

Creates a backup before modifying.

Usage:
    python fix_master_columns.py                    # dry-run (preview)
    python fix_master_columns.py --apply            # apply fixes
"""
import os
import sys
import shutil
import argparse
from datetime import datetime
from openpyxl import load_workbook

MASTER = './output/master.xlsx'

OLD_HEADERS = ['ФИО', 'Дата рождения', '№ полиса', 'Начало обслуживания',
               'Конец обслуживания', 'Страховая компания', 'Страхователь',
               'Источник файла', 'Дата обработки']

NEW_HEADERS = ['ФИО', 'Дата рождения', '№ полиса', 'Начало обслуживания',
               'Конец обслуживания', 'Страховая компания', 'Страхователь',
               'Клиника', 'Комментарий в полис', 'Источник файла', 'Дата обработки']

PHANTOM_SOURCE = '2042_records_2026-03-19.xlsx'


def fix(apply: bool):
    if not os.path.exists(MASTER):
        print(f"ERROR: {MASTER} not found")
        sys.exit(1)

    wb = load_workbook(MASTER)
    ws = wb['Данные']

    # Read current headers
    max_col = ws.max_column
    headers = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
    print(f"Current headers ({max_col} cols): {headers}")
    print(f"Total rows (incl header): {ws.max_row}")

    # Identify row types
    old_rows = []       # rows with old 9-col layout
    new_rows = []       # rows with new 11-col layout (written today)
    phantom_rows = []   # re-ingested from own report

    for r in range(2, ws.max_row + 1):
        # Check if col 10 has data (new-layout rows have Источник файла in col 10)
        col10 = ws.cell(row=r, column=10).value
        col8 = ws.cell(row=r, column=8).value

        if col10 is not None:
            # New layout: cols 8=Клиника, 9=Комментарий, 10=Источник, 11=Дата
            source = str(ws.cell(row=r, column=10).value or '')
            if PHANTOM_SOURCE in source:
                phantom_rows.append(r)
            else:
                new_rows.append(r)
        else:
            # Old layout: cols 8=Источник, 9=Дата
            old_rows.append(r)

    print(f"\nOld-layout rows: {len(old_rows)}")
    print(f"New-layout rows (today, real): {len(new_rows)}")
    print(f"Phantom rows (re-ingested report): {len(phantom_rows)}")

    if phantom_rows:
        print(f"\nPhantom rows to DELETE (source={PHANTOM_SOURCE}):")
        for r in phantom_rows[:5]:
            fio = ws.cell(row=r, column=1).value
            print(f"  row {r}: {fio}")
        if len(phantom_rows) > 5:
            print(f"  ... and {len(phantom_rows) - 5} more")

    if not apply:
        print("\nDRY RUN — no changes made. Run with --apply to fix.")
        wb.close()
        return

    # Backup
    backup = MASTER.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    shutil.copy2(MASTER, backup)
    print(f"\nBackup saved: {backup}")

    # Step 1: Delete phantom rows (bottom-up to preserve indices)
    for r in sorted(phantom_rows, reverse=True):
        ws.delete_rows(r, 1)
    print(f"Deleted {len(phantom_rows)} phantom rows")

    # After deletion, row indices shifted. Re-scan to find old vs new rows.
    old_rows_2 = []
    new_rows_2 = []
    for r in range(2, ws.max_row + 1):
        col10 = ws.cell(row=r, column=10).value
        if col10 is not None:
            new_rows_2.append(r)
        else:
            old_rows_2.append(r)

    # Step 2: Fix old-layout rows — shift col 8-9 → col 10-11, clear 8-9
    for r in old_rows_2:
        src_file = ws.cell(row=r, column=8).value  # Источник файла
        date_proc = ws.cell(row=r, column=9).value  # Дата обработки
        ws.cell(row=r, column=10, value=src_file)
        ws.cell(row=r, column=11, value=date_proc)
        ws.cell(row=r, column=8, value='')   # Клиника (unknown for old data)
        ws.cell(row=r, column=9, value='')   # Комментарий в полис
    print(f"Fixed {len(old_rows_2)} old-layout rows (shifted Источник/Дата → cols 10-11)")

    # Step 3: New-layout rows are already correct (cols 8-11 in right order)
    print(f"New-layout rows OK: {len(new_rows_2)} rows (no changes needed)")

    # Step 4: Fix headers
    for col_idx, name in enumerate(NEW_HEADERS, 1):
        ws.cell(row=1, column=col_idx, value=name)
    # Clear any extra None headers
    for c in range(len(NEW_HEADERS) + 1, max_col + 1):
        ws.cell(row=1, column=c, value=None)
    print(f"Updated headers to: {NEW_HEADERS}")

    wb.save(MASTER)
    wb.close()
    print(f"\nDone! Fixed {MASTER}")
    print(f"  Old rows: Клиника/Комментарий set to empty (will populate on next match)")
    print(f"  Phantom rows: removed")
    print(f"  Headers: migrated to 11-column layout")


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--apply', action='store_true', help='Actually apply fixes (default is dry-run)')
    args = parser.parse_args()
    fix(args.apply)
