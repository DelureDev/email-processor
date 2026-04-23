"""Per-row Франшиза comment extraction for Alfa parser (v1.10.13)."""
from openpyxl import Workbook
from parsers.alfa import parse


def _write_alfa_xlsx(path, data_rows, *, header_row_idx=6, include_comment_col=True):
    """Build a minimal Alfa-format xlsx.

    data_rows: list of (num, polis, fio, birth, start, end, comment) tuples.
    include_comment_col: if False, omit the 'Вид медицинского обслуживания' header
    """
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value='АО "АльфаСтрахование"')
    ws.cell(row=4, column=1, value='Список застрахованных')
    ws.cell(row=5, column=1, value='для обслуживания в сети клиник ________')
    header_row = header_row_idx + 1  # openpyxl is 1-indexed
    ws.cell(row=header_row, column=1, value='№ п/п')
    ws.cell(row=header_row, column=2, value='№ полиса')
    ws.cell(row=header_row, column=3, value='ФИО')
    ws.cell(row=header_row, column=4, value='Дата рождения')
    ws.cell(row=header_row, column=5, value='Адрес')
    ws.cell(row=header_row, column=6, value='Группа, № договора')
    ws.cell(row=header_row, column=7, value='Период с')
    ws.cell(row=header_row, column=8, value='Период по')
    if include_comment_col:
        ws.cell(row=header_row, column=9, value='Вид медицинского обслуживания')
    for i, (num, polis, fio, birth, start, end, comment) in enumerate(data_rows, start=header_row + 1):
        ws.cell(row=i, column=1, value=num)
        ws.cell(row=i, column=2, value=polis)
        ws.cell(row=i, column=3, value=fio)
        ws.cell(row=i, column=4, value=birth)
        ws.cell(row=i, column=6, value='Группа X; №12345; ООО "ТестСтрах"')
        ws.cell(row=i, column=7, value=start)
        ws.cell(row=i, column=8, value=end)
        if include_comment_col and comment is not None:
            ws.cell(row=i, column=9, value=comment)
    wb.save(path)


def test_alfa_extracts_franchise_per_row(tmp_path):
    """Different Франшиза values per-row land in each record's Комментарий в полис."""
    path = tmp_path / "alfa_multi.xlsx"
    _write_alfa_xlsx(path, [
        (1, 'POL001', 'ИВАНОВ И И', '01.01.1990', '01.01.2026', '31.12.2026',
         'Франшиза 20%: амбулаторно-поликлиническое'),
        (2, 'POL002', 'ПЕТРОВ П П', '02.02.1985', '01.01.2026', '31.12.2026',
         'Франшиза 45%: амбулаторно + стационар'),
        (3, 'POL003', 'СИДОРОВ С С', '03.03.1980', '01.01.2026', '31.12.2026',
         'Без франшизы, амбулаторно'),
    ])

    records = parse(str(path))

    assert len(records) == 3
    assert records[0]['Комментарий в полис'] == 'Франшиза 20%: амбулаторно-поликлиническое'
    assert records[1]['Комментарий в полис'] == 'Франшиза 45%: амбулаторно + стационар'
    # "Без франшизы" — "франшизы" (genitive, ending 'ы') does NOT contain "франшиза" (ending 'а').
    # Parser does not set the field; downstream clinic_matcher may fill it instead.
    assert not records[2].get('Комментарий в полис')


def test_alfa_no_comment_column_no_extraction(tmp_path):
    """If no comment column header, parser does not set Комментарий в полис."""
    path = tmp_path / "alfa_no_comment_col.xlsx"
    _write_alfa_xlsx(path, [
        (1, 'POL001', 'ИВАНОВ И И', '01.01.1990', '01.01.2026', '31.12.2026', None),
    ], include_comment_col=False)

    records = parse(str(path))

    assert len(records) == 1
    assert not records[0].get('Комментарий в полис')


def test_alfa_franchise_cell_uppercase(tmp_path):
    """Franchise keyword match is case-insensitive; full original text preserved."""
    path = tmp_path / "alfa_upper.xlsx"
    _write_alfa_xlsx(path, [
        (1, 'POL001', 'ИВАНОВ И И', '01.01.1990', '01.01.2026', '31.12.2026',
         'ФРАНШИЗА 30%: амбулаторно'),
    ])

    records = parse(str(path))

    assert len(records) == 1
    assert records[0]['Комментарий в полис'] == 'ФРАНШИЗА 30%: амбулаторно'


def test_alfa_non_franchise_comment_not_extracted(tmp_path):
    """A comment cell without 'франшиза' stays unset at parser level."""
    path = tmp_path / "alfa_no_franchise.xlsx"
    _write_alfa_xlsx(path, [
        (1, 'POL001', 'ИВАНОВ И И', '01.01.1990', '01.01.2026', '31.12.2026',
         'Амбулаторно-поликлиническое обслуживание без стоматологии'),
    ])

    records = parse(str(path))

    assert len(records) == 1
    assert not records[0].get('Комментарий в полис')


class TestMainCommentPrecedence:
    """Parser-set Комментарий в полис wins over clinic_matcher fallback."""

    def test_parser_franchise_wins_over_clinic_matcher(self, tmp_path, monkeypatch):
        """Parser's per-row Франшиза beats clinic_matcher's file-level comment."""
        import main

        path = tmp_path / "alfa.xlsx"
        _write_alfa_xlsx(path, [
            (1, 'POL001', 'ИВАНОВ И И', '01.01.1990', '01.01.2026', '31.12.2026',
             'Франшиза 55%: per-row value'),
        ])

        monkeypatch.setattr(main, 'detect_clinic',
                            lambda *a, **kw: ('Клиника1', True, '123'))
        monkeypatch.setattr(main, 'extract_policy_comment',
                            lambda *a, **kw: 'FILE-LEVEL-CLINIC-COMMENT')

        stats = main.make_stats()
        stats['master_path'] = str(tmp_path / 'master.xlsx')
        config = {'processing': {'deduplicate': False}}

        main.process_file(str(path), stats['master_path'], config, stats,
                         existing_keys=set(), dry_run=True, pending=None)

        assert len(stats['new_records']) == 1
        assert stats['new_records'][0]['Комментарий в полис'] == 'Франшиза 55%: per-row value'


    def test_clinic_matcher_fills_when_parser_empty(self, tmp_path, monkeypatch):
        """When parser sets no Комментарий, clinic_matcher's fallback populates it."""
        import main

        path = tmp_path / "alfa_no_franchise.xlsx"
        _write_alfa_xlsx(path, [
            (1, 'POL001', 'ИВАНОВ И И', '01.01.1990', '01.01.2026', '31.12.2026',
             'Амбулаторно-поликлиническое без стоматологии'),
        ])

        monkeypatch.setattr(main, 'detect_clinic',
                            lambda *a, **kw: ('Клиника1', True, '123'))
        monkeypatch.setattr(main, 'extract_policy_comment',
                            lambda *a, **kw: 'CLINIC-MATCHER-FALLBACK')

        stats = main.make_stats()
        stats['master_path'] = str(tmp_path / 'master.xlsx')
        config = {'processing': {'deduplicate': False}}

        main.process_file(str(path), stats['master_path'], config, stats,
                         existing_keys=set(), dry_run=True, pending=None)

        assert len(stats['new_records']) == 1
        assert stats['new_records'][0]['Комментарий в полис'] == 'CLINIC-MATCHER-FALLBACK'
