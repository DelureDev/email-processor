#!/usr/bin/env python3
"""One-time: dedup master files + fix network share CSVs."""
import openpyxl
import csv
import os
import pandas as pd
from main import load_config
from writer import COLUMNS, _safe
from parsers.utils import norm_date_pad

config = load_config()
share = config.get('output', {}).get('csv_export_folder', '').strip()

# 1. Dedup master.xlsx
path = './output/master.xlsx'
wb = openpyxl.load_workbook(path)
ws = wb['Данные']
seen = set()
dupes = []
for r in range(2, ws.max_row + 1):
    key = tuple(str(ws.cell(row=r, column=c).value or '') for c in range(1, 12))
    if key in seen:
        dupes.append(r)
    seen.add(key)
for r in reversed(dupes):
    ws.delete_rows(r, 1)
wb.save(path)
wb.close()
print(f'master.xlsx: removed {len(dupes)} duplicates, {ws.max_row - 1} rows remain')

# 2. Dedup master.csv
csv_path = './output/master.csv'
if os.path.exists(csv_path):
    with open(csv_path, 'r', encoding='utf-8-sig', newline='') as f:
        rows = list(csv.reader(f, delimiter=';'))
    header = rows[0] if rows else []
    seen_csv = set()
    kept = [header]
    removed = 0
    for row in rows[1:]:
        key = tuple(row)
        if key in seen_csv:
            removed += 1
        else:
            seen_csv.add(key)
            kept.append(row)
    with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.writer(f, delimiter=';', lineterminator='\r\n')
        w.writerows(kept)
    print(f'master.csv: removed {removed} duplicates, {len(kept) - 1} rows remain')

# 3. Fix network share CSVs
if not share:
    print('No csv_export_folder configured, skipping share fix')
    exit(0)

if not os.path.isdir(share):
    print(f'Share not reachable: {share}')
    exit(1)

# Network CSV columns (with ID Клиники after Клиника)
csv_columns = []
for c in COLUMNS:
    csv_columns.append(c)
    if c == 'Клиника':
        csv_columns.append('ID Клиники')

# Read today's records from master
df = pd.read_excel(path, dtype=str).fillna('')
today_mask = df['Дата обработки'].map(lambda s: norm_date_pad(str(s))) == '20.03.2026'
today_records = df[today_mask].to_dict('records')
march_mask = df['Дата обработки'].map(lambda s: norm_date_pad(str(s)).endswith('03.2026'))
march_records = df[march_mask].to_dict('records')

# 3a. Rewrite daily CSV from scratch
daily = os.path.join(share, 'records_2026-03-20.csv')
with open(daily, 'w', encoding='utf-8-sig', newline='') as f:
    w = csv.DictWriter(f, fieldnames=csv_columns, extrasaction='ignore', delimiter=';', lineterminator='\r\n')
    w.writeheader()
    for rec in today_records:
        w.writerow({k: _safe(v) for k, v in rec.items()})
print(f'Daily CSV: wrote {len(today_records)} records to {daily}')

# 3b. Rewrite monthly master CSV from scratch (with ID Клиники column)
monthly = os.path.join(share, 'master_2026-03.csv')
with open(monthly, 'w', encoding='utf-8-sig', newline='') as f:
    w = csv.DictWriter(f, fieldnames=csv_columns, extrasaction='ignore', delimiter=';', lineterminator='\r\n')
    w.writeheader()
    for rec in march_records:
        w.writerow({k: _safe(v) for k, v in rec.items()})
print(f'Monthly CSV: wrote {len(march_records)} records to {monthly}')
